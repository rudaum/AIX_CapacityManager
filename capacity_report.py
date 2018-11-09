#!/usr/bin/python
# - Purpose:
#       To convert range of dates
# - Author:
#       Rudolf Wolter
# - Contact for questions and/or comments:
#       rudolf.wolter@kuehne-nagel.com
# - Parameters:
#       <accepted arguments>
# - Version Releases and modifications.
#       <versions history log>

### START OF MODULE IMPORTS
# --------------------------------------------------------------- #
from os import path
from lib.dblib import Vmstat, Lparstat
from datetime import datetime, timedelta
from argparse import ArgumentParser, RawTextHelpFormatter
from openpyxl import Workbook
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.text import RichTextProperties, Paragraph, \
    ParagraphProperties, CharacterProperties, PatternFillProperties
from openpyxl.chart import Series, LineChart, Reference, AreaChart
from openpyxl.chart.text import RichText
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.trendline import Trendline, GraphicalProperties
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
# --- For Transparency
from openpyxl.compat import unicode
from openpyxl.descriptors import Typed
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors.nested import NestedInteger
from openpyxl.xml.constants import DRAWING_NS

# --------------------------------------------------------------- #
### END OF MODULE IMPORTS

### START OF GLOBAL VARIABLES DECLARATION
# --------------------------------------------------------------- #
# --- Date / Time Variables
PEAKSTART = '08:00'
PEAKSTOP = '20:00'
TRESHOLDW = 20
TRESHOLDC = 8
FORECAST = 0.33
# ---

VMSTATLOCALDIR = path.dirname(path.realpath(__file__)) + '/reports'
WORKBOOKNAME = path.dirname(path.realpath(__file__)) + '/SomethingWentBad.xlsx'
BL_IBROKER = ['sindbad', 'ibaz', 'ibedi1', 'ibedi2',
              'ibedi3', 'ibedi4', 'ibcom1', 'ibcom2',
              'ibcom3', 'parisade', 'ibcomtest',
              'salomo', 'alibaba', 'maruf', 'dehamsp41']
BL_FTP = ['ftpext', 'ftpint']
BL_QMM = ['dehamsp34', 'denotsp34']
BL_WORKPLACE = ['kn-workplace', 'kn-workplace-db']
BASELINE = BL_IBROKER + BL_FTP + BL_QMM + BL_WORKPLACE

BGCOLOR = '8CB4E2'
FONTCOLOR = '000000'
FONTNAME = 'Arial Unicode MS'
BORDER = Border(
    left=Side(border_style='double'),
    right=Side(border_style='double'),
    top=Side(border_style='double'),
    bottom=Side(border_style='double')
)


# --------------------------------------------------------------- #
### END OF GLOBAL VARIABLES DECLARATION\

### START OF FUNCTIONS DECLARATION
# --------------------------------------------------------------- #
def parse_args():
    """
    Purpose:
        To parse the scripts given arguments and to generate a dictionary with the values.
    Returns:
        options: A Dictionary with all given, validated arguments.
    Parameters:
    """
    parser = ArgumentParser(description='Generates an Excel with Charts of Capacity Usage and Trends.',
                            formatter_class=RawTextHelpFormatter)

    # Declaring Arguments
    parser.add_argument("-s", "--server", dest="server", required=True,
                        help="The server name or baseline to generate the reports. Accepted values are:\n"
                             "- all: to generate reports for all servers.\n"
                             "- servername: o generate reports for a particular server.\n"
                             "- ftp: to generate reports for FTP servers.\n"
                             "- ibroker: to generate reports for Information Broker servers.\n"
                             "- qmm: to generate reports for QMM servers.\n"
                             "- workplace: to generate reports for WORKPLACE servers.\n"
                        )
    parser.add_argument("-r", "--range", dest="range", required=True,
                        help="ALL for all available dates, or a specific time range.\n"
                             "Specific day range format: dd/mm/yyyy-dd/mm/yyyy")
    parser.add_argument("-d", "--dont-skip", dest="dontskip", required=False, default=False, action='store_true',
                        help="Don't skip weekends when generating the reports.")
    parser.add_argument("-f", "--full-day", dest="fullday", required=False, default=False, action='store_true',
                        help="Use full day statistics, instead of Business Peak hours.")

    args = vars(parser.parse_args())  # A dictionary with arguments

    # Lowering string
    args['server'] = args['server'].lower()
    args['range'] = args['range'].lower()
    args['range'] = args['range'].replace('.', '/')

    # Setting workbook name
    _tr = 'alltime' if args['range'] == 'all' else args['range'].replace('/', '')
    args['workbookname'] = \
        '{}/{}_{}_perfreport.xlsx' \
        .format(path.dirname(path.realpath(__file__)), _tr, args['server'])

    # Translating Server Baseline
    if args['server'] == 'all':
        args['server'] = BASELINE
    elif args['server'] == 'ibroker':
        args['server'] = BL_IBROKER
    elif args['server'] == 'ftp':
        args['server'] = BL_FTP
    elif args['server'] == 'qmm':
        args['server'] = BL_QMM
    elif args['server'] == 'workplace':
        args['server'] = BL_WORKPLACE
    else:
        args['server'] = eval("['{}']".format(args['server']))

    # Server is not in the baseline
    if not set(BASELINE).issuperset(set(args['server'])):
        print('Unknown Server! Server must be one of the following:\n{}\n'.format(BASELINE))
        exit(2)

    # Range argument is shorter than 3 characters
    if len(args['range']) < 3:
        print('Provide at least 3 characters for the date range. \n')
        exit(2)

    # Checking if range is 'all' or in 'dd/mm/yyyy-dd/mm/yyyy' format
    if args['range'] != 'all' and (args['range'].count('/') != 4 or args['range'].count('-') != 1):
        print args['range']
        print('Invalid date range. Use "ALL" or the date format "dd/mm/yyyy-dd/mm/yyyy". \n')
        exit(2)

    if args['range'] != 'all':
        d8begin = ''
        d8end = ''

        # Validating the Start date from the range
        try:
            d8begin = datetime.strptime(args['range'].split('-')[0], '%d/%m/%Y')
        except ValueError:
            print('Invalid START date. Use "ALL" or the date format "dd/mm/yyyy-dd/mm/yyyy". \n')
            exit(2)

        # Validating the END date from the range
        try:
            d8end = datetime.strptime(args['range'].split('-')[1], '%d/%m/%Y')
        except ValueError:
            print('Invalid END date. Use "ALL" or the date format "dd/mm/yyyy-dd/mm/yyyy". \n')
            exit(2)

        # Setting Start and End dates
        if d8begin >= d8end:
            print('START date is the same or after END date \n')
            exit(2)
        else:
            args['startdate'] = d8begin
            args['enddate'] = d8end

    else:
        args['startdate'] = datetime(1970, 01, 01)
        args['enddate'] = datetime(2999, 12, 31)

    return args
# --------------------------------------------------------------- #
# --------------------------------------------------------------- #
def mk_readme_sheet(wb):
    """
    Purpose:
        To create a README sheet with useful information about the other graphs.

    Parameters:
        wb - The current Workbook
    """
    ws = wb.worksheets[0]
    ws.title = 'README'
    # "Coloring" the sheet's background, to hide the data.
    area = Reference(ws, min_col=1, min_row=1, max_row=60, max_col=40)
    for cell in area.cells:
        ws[cell].fill = PatternFill(bgColor=BGCOLOR, fgColor=BGCOLOR, fill_type="solid")
        ws[cell].font = Font(name=FONTNAME, color=FONTCOLOR, size=12)

    # Setting Columns size ---
    ws.column_dimensions['B'].width = 96
    # ---

    # Information Block ---
    ws['B2'].font = Font(name="Arial Black", size=14)
    ws['B2'].value = 'About This:'
    ws['B3'].border = BORDER
    ws['B3'].alignment = Alignment(wrapText=True)
    ws['B3'].value = '    The Charts are generated from daily averages of 10-seconds interval snapshot samples, ' \
                     'that were collected during week days and during Peak Business Hours ({} - {} CEST). ' \
                     'Each sheet contains Charts for one or more Servers. ' \
                     'Note that weekends will not be included in the charts.'.format(PEAKSTART, PEAKSTOP)
    # ---

    # CPU_Relative Block ---
    ws['B5'].font = Font(name="Arial Black", size=14)
    ws['B5'].value = 'CPU_Relative Explained:'
    ws['B6'].border = BORDER
    ws['B6'].alignment = Alignment(wrapText=True)
    ws['B6'].value = "    The Charts represents the daily averages of 4 measures:\n" \
                     "- 'Average Busy': The average of the CPU utilisation of that day. ex. 60 would mean that " \
                     "the server had an average of 60% of CPU utilisation during business hours for that day.\n" \
                     "- 'Linear(Average Busy)': A Linear trend generated from the Average Busy of all days," \
                     " including a prognosis for an extra of 33% of the selected date range (in days).\n" \
                     "- '% of Samples above {}% Cpu Usage': Represents an average the percentage of Samples collected" \
                     " of which the CPU utilisation were above {}%.\n" \
                     "- '% of Samples above {}% Cpu Usage': Represents an average the percentage of Samples collected" \
                     " of which the CPU utilisation were above {}%, it doesn't count the {}% samples." \
        .format(100 - TRESHOLDW, 100 - TRESHOLDW, 100 - TRESHOLDC, 100 - TRESHOLDC, 100 - TRESHOLDW)
    # ---

    # CPU_Relative Block ---
    ws['B8'].font = Font(name="Arial Black", size=14)
    ws['B8'].value = 'CPU_Absolute Explained:'
    ws['B9'].border = BORDER
    ws['B9'].alignment = Alignment(wrapText=True)
    ws['B9'].value = "    The Charts represent the real utilisation of Physical cores in comparison to the server's" \
                     " virtual processors. The real physical utilization can not be bigger than the virtual procs," \
                     " therefore if 'Average Busy' area is reaching the top of the chart, then it means the server's" \
                     " load is  reaching its Maximum capacity. There is also a linear trend indicator for that date " \
                     "interval, that includes a prognosis for an extra of 33% of the selected date range (in days). \n"
    # ---
# --------------------------------------------------------------- #

# --------------------------------------------------------------- #
def mk_cpurelative_sheet(wb, _options):
    """
    Purpose:
        To create the Workbook(excel file) with Relative CPU usage graphs using the data gathered from vmstat

    Parameters:
        wb - The current Workbook
        _options - the arguments passed to the script
    """
    # --- 'CPU' Sheet
    ws = wb.create_sheet('CPU_Relative')
    looprow_marker = 1  # Keeps a track of the next row that a data record will be inserted
    loopcol = 40  # increases in 4 columns for each server
    graphpos = 2  # Which row to place the First Graph

    # "Coloring" the sheet's background, to hide the data.
    area = Reference(ws, min_col=1, min_row=1, max_row=500, max_col=40)
    for cell in area.cells:
        ws[cell].fill = PatternFill(bgColor=BGCOLOR, fgColor=BGCOLOR, fill_type="solid")
        ws[cell].font = Font(color=BGCOLOR)

    for server in _options['server']:
        print('Generating "vmstat" reports for {} ...'.format(server))
        looprow = looprow_marker
        ws.cell(row=1, column=loopcol, value='Date')
        ws.cell(row=1, column=loopcol + 1, value='Average Busy')
        ws.cell(row=1, column=loopcol + 2, value='% of Samples above {}% Cpu Usage'.format(100 - TRESHOLDW))
        ws.cell(row=1, column=loopcol + 3, value='% of Samples above {}% Cpu Usage'.format(100 - TRESHOLDC))

        # Querying Server and its data from Database
        server_entries = Vmstat().query_by('servername', server)

        # Looping over server's entries
        for entry in server_entries:
            date = entry.date
            average = entry.peak_avg_busy
            warning = entry.peak_avg_warning
            critical = entry.peak_avg_critical

            # if the entry's date is newer than range's start and older then range's stop
            if _options['startdate'] <= date <= _options['enddate']:
                # if we want to skip weekend days
                if not _options['dontskip'] and date.weekday() < 5:
                    looprow = looprow + 1
                    ws.cell(row=looprow, column=loopcol, value=date)
                    ws.cell(row=looprow, column=loopcol + 1, value=average)
                    ws.cell(row=looprow, column=loopcol + 2, value=warning)
                    ws.cell(row=looprow, column=loopcol + 3, value=critical)

                # if we want to include weekends
                elif _options['dontskip']:
                    looprow = looprow + 1
                    ws.cell(row=looprow, column=loopcol, value=date)
                    ws.cell(row=looprow, column=loopcol + 1, value=average)
                    ws.cell(row=looprow, column=loopcol + 2, value=warning)
                    ws.cell(row=looprow, column=loopcol + 3, value=critical)

        try:
            # --- Setting Chart Properties
            chart = LineChart()
            chart.title = '{} %CPU'.format(server)
            chart.style = 3
            chart.x_axis.number_format = 'dd/mm'
            chart.x_axis.majorTimeUnit = "days"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 100

            # --- All this to rotate the 'date' axis in 270 degrees
            rot = RichTextProperties(vert='vert270')
            axis = CharacterProperties()
            chart.x_axis.textProperties = \
                RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)
            # ---
            chart.height = 10
            chart.width = 25
            chart.legend.position = "t"
            chart.legend.layout = Layout(manualLayout=ManualLayout(yMode='edge', xMode='edge', h=0.1, w=0.8))
            chart.plot_area.layout = Layout(manualLayout=ManualLayout(yMode='edge', xMode='edge'))
            pf = PatternFillProperties(prst='pct25', fgClr=ColorChoice(srgbClr='8FAF52'))
            chart.plot_area.graphicalProperties = GraphicalProperties(pattFill=pf)

            # --- Looping over Columns to build the Chart
            # columns Average, Warning and Critical
            for col in range(loopcol + 1, loopcol + 4):
                values = Reference(ws, min_col=col, min_row=1, max_row=looprow)
                series = Series(values, title_from_data=True)  # creating the Serie
                chart.series.append(series)

            # --- Setting Plot Area Format
            # Formatting Series
            chart.series[0].graphicalProperties.line = LineProperties(solidFill=ColorChoice(prstClr='green'))
            chart.series[1].graphicalProperties.line = LineProperties(solidFill=ColorChoice(prstClr='orange'))
            chart.series[2].graphicalProperties.line = LineProperties(solidFill=ColorChoice(prstClr='red'))

            # Creating and Formatting Trend Line
            chart.series[0].trendline = Trendline(forward=str(1 + int(looprow * FORECAST)))
            chart.series[0].trendline.graphicalProperties = GraphicalProperties()
            chart.series[0].trendline.graphicalProperties.line.solidFill = ColorChoice(prstClr='purple')
            chart.series[0].trendline.graphicalProperties.line.width = 25000

            # Setting the 'date' x-axis, with forecast of 33% of the time range
            i = 0
            for i in range(looprow, 2 + int(looprow * (1 + FORECAST))):
                ws.cell(row=i, column=loopcol).value = \
                    ws.cell(row=i - 1, column=loopcol).value + timedelta(days=1)
            dates = Reference(ws, min_col=loopcol, min_row=2, max_row=i)
            chart.set_categories(dates)

            # Set the Starting column for the next server
            loopcol = loopcol + 4

            # Adding the Chart
            ws.add_chart(chart, "A{}".format(graphpos))

            # Adding The Comments Session
            # Setting Analysis Column size ---
            ws.column_dimensions['P'].width = 75
            ws.merge_cells('P{}:P{}'.format(graphpos + 1, graphpos + 17))

            ws['P{}'.format(graphpos)].font = Font(name=FONTNAME, size=14, color=FONTCOLOR)
            ws['P{}'.format(graphpos)].value = '{} Analysis:'.format(server)
            ws['P{}'.format(graphpos + 1)].alignment = Alignment(horizontal='left', vertical='top', wrapText=True)

            area = Reference(ws, min_col=16, min_row=graphpos+1, max_row=graphpos+17, max_col=16)
            for cell in area.cells:
                ws[cell].font = Font(name=FONTNAME, size=11, color=FONTCOLOR)
                ws[cell].fill = PatternFill(fill_type="solid", start_color='FFFFFF', end_color='FFFFFF')
                ws[cell].border = BORDER

            # Updating the Graphic Positioner
            graphpos = graphpos + 19

        except ValueError as err_msg:
            print 'Error while processing vmstat data for server {}! Could not create the Chart.\n' \
                  'Extend error message: \n' \
                  '{}'.format(server, err_msg)
# --------------------------------------------------------------- #

# --------------------------------------------------------------- #
def mk_cpuabsolute_sheet(wb, _options):
    """
    Purpose:
        To create the Workbook(excel file) with Relative CPU usage graphs using the data gathered from vmstat

    Parameters:
        wb - The current Workbook
        _options - the arguments passed to the script
    """
    # --- 'CPU' Sheet
    ws = wb.create_sheet('CPU_Absolute')
    looprow_marker = 1  # Keeps a track of the next row that a data record will be inserted
    loopcol = 40  # increases in 4 columns for each server
    graphpos = 2  # Which row to place the First Graph

    # "Coloring" the sheet's background, to hide the data.
    area = Reference(ws, min_col=1, min_row=1, max_row=500, max_col=40)
    for cell in area.cells:
        ws[cell].fill = PatternFill(bgColor=BGCOLOR, fgColor=BGCOLOR, fill_type="solid")
        ws[cell].font = Font(color=BGCOLOR)

    for server in _options['server']:
        print('Generating "lparstat" reports for {} ...'.format(server))
        cpu_type = cpu_mode = ''
        vprocs = 0
        looprow = looprow_marker
        ws.cell(row=1, column=loopcol, value='Date')
        ws.cell(row=1, column=loopcol + 3, value='Average Busy')
        ws.cell(row=1, column=loopcol + 2, value='Entitled capacity')
        ws.cell(row=1, column=loopcol + 1, value='Virtual Procs')

        # Querying Server and its data from Database
        server_entries = Lparstat().query_by('servername', server)
        highest_vprocs = 0
        physc_usage_list = []

        for entry in server_entries:
            date = entry.date
            cpu_type = entry.cpu_type
            cpu_mode = entry.cpu_mode
            vprocs = entry.vprocs

            if vprocs > highest_vprocs:
                highest_vprocs = vprocs

            ent_cap = entry.ent_cap
            physc_peak_average = entry.peak_avg_physc

            # if the entry's date is newer than range's start and older then range's stop
            if _options['startdate'] <= date <= _options['enddate']:
                # if we want to skip weekend days
                if not _options['dontskip'] and date.weekday() < 5:
                    looprow = looprow + 1
                    ws.cell(row=looprow, column=loopcol, value=date)
                    ws.cell(row=looprow, column=loopcol + 3, value=physc_peak_average)
                    ws.cell(row=looprow, column=loopcol + 2, value=ent_cap)
                    ws.cell(row=looprow, column=loopcol + 1, value=vprocs)
                    physc_usage_list.append(physc_peak_average)

                # if we want to include weekends
                elif _options['dontskip']:
                    looprow = looprow + 1
                    ws.cell(row=looprow, column=loopcol, value=date)
                    ws.cell(row=looprow, column=loopcol + 3, value=physc_peak_average)
                    ws.cell(row=looprow, column=loopcol + 2, value=ent_cap)
                    ws.cell(row=looprow, column=loopcol + 1, value=vprocs)

        try:
            # Setting the FORECAST data
            for i in range(looprow, 2 + int(looprow * (1 + FORECAST))):
                ws.cell(row=i, column=loopcol).value = ws.cell(row=i - 1, column=loopcol).value + timedelta(days=1)
                ws.cell(row=i, column=loopcol + 1).value = ws.cell(row=i - 1, column=loopcol + 1).value
                ws.cell(row=i, column=loopcol + 2).value = ws.cell(row=i - 1, column=loopcol + 2).value

            # --- Setting Chart Properties
            chart = AreaChart()
            chart.title = '{} - Physical Cores ({} {}/{} Virt. Procs)' \
                .format(server, vprocs, cpu_type, cpu_mode)
            chart.style = 3
            chart.x_axis.number_format = 'dd/mm'
            chart.x_axis.majorTimeUnit = "days"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = highest_vprocs

            # --- All this to rotate the 'date' axis in 270 degrees
            rot = RichTextProperties(vert='vert270')
            axis = CharacterProperties()
            chart.x_axis.textProperties = \
                RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)
            # ---
            chart.height = 10
            chart.width = 25
            chart.legend.position = "t"
            chart.legend.layout = Layout(manualLayout=ManualLayout(yMode='edge', xMode='edge', h=0.1, w=0.8))
            chart.plot_area.layout = Layout(manualLayout=ManualLayout(yMode='edge', xMode='edge'))
            pf = PatternFillProperties(prst='pct20', fgClr=ColorChoice(srgbClr='8FAF52'))
            chart.plot_area.graphicalProperties = GraphicalProperties(pattFill=pf)

            # --- Creating and formatting the Chart's Series
            # Virtual Procs Series
            values = Reference(ws, min_col=loopcol + 1, min_row=1, max_row=1 + int(looprow * (1 + FORECAST)))
            serie = Series(values, title_from_data=True)  # creating the Serie
            pf = PatternFillProperties(prst='smGrid', fgClr=ColorChoice(srgbClr='D4ECBA'))
            serie.graphicalProperties = GraphicalProperties(pattFill=pf)
            serie.graphicalProperties.line.width = 25000
            chart.series.append(serie)

            # Formatting Entitled Capacity Series
            values = Reference(ws, min_col=loopcol + 2, min_row=1, max_row=1 + int(looprow * (1 + FORECAST)))
            serie = Series(values, title_from_data=True)  # creating the Serie
            serie.graphicalProperties = GraphicalProperties(solidFill='D4ECBA')
            serie.graphicalProperties.line.width = 25000
            chart.series.append(serie)

            # Formatting Physc Busy Series
            values = Reference(ws, min_col=loopcol + 3, min_row=1, max_row=looprow)
            serie = Series(values, title_from_data=True)  # creating the Serie
            serie.graphicalProperties = GraphicalProperties()
            serie.graphicalProperties.line.solidFill = '558ED5'
            serie.graphicalProperties.solidFill = MyColorChoice(srgbClr=MyRGBColor('558ED5', alpha=35000))
            serie.graphicalProperties.line.width = 25000
            chart.series.append(serie)

            # Creating and Formatting Trend Line
            chart.series[2].trendline = Trendline(forward=str(1 + int(looprow * FORECAST)))
            chart.series[2].trendline.graphicalProperties = GraphicalProperties()
            chart.series[2].trendline.graphicalProperties.line.solidFill = ColorChoice(prstClr='purple')
            chart.series[2].trendline.graphicalProperties.line.width = 25000

            # Setting the 'date' x-axis, with forecast of 33% of the time range
            dates = Reference(ws, min_col=loopcol, min_row=2, max_row=1 + int(looprow * (1 + FORECAST)))
            chart.set_categories(dates)

            # Set the Starting column for the next server
            loopcol = loopcol + 4

            # Adding the Chart
            ws.add_chart(chart, "A{}".format(graphpos))

            # Adding The Comments Session
            # Setting Analysis Column size ---
            ws.column_dimensions['P'].width = 75
            ws.merge_cells('P{}:P{}'.format(graphpos + 1, graphpos + 17))

            ws['P{}'.format(graphpos)].font = Font(name=FONTNAME, size=14, color=FONTCOLOR)
            ws['P{}'.format(graphpos)].value = '{} Analysis:'.format(server)
            ws['P{}'.format(graphpos + 1)].alignment = Alignment(horizontal='left', vertical='top', wrapText=True)

            area = Reference(ws, min_col=16, min_row=graphpos + 1, max_row=graphpos + 17, max_col=16)
            for cell in area.cells:
                ws[cell].font = Font(name=FONTNAME, size=11, color=FONTCOLOR)
                ws[cell].fill = PatternFill(fill_type="solid", start_color='FFFFFF', end_color='FFFFFF')
                ws[cell].border = BORDER

            # Updating the Graphic Positioner
            graphpos = graphpos + 19
        except (ValueError, TypeError) as err_msg:
            print 'Error while processing lparstat data for server {}! Could not create the Chart.\n' \
                  'Extend error message: \n' \
                  '{}'.format(server, err_msg)




# --------------------------------------------------------------- #
### END OF FUNCTIONS DECLARATION

### START OF CLASS DEFINITIONS
# --------------------------------------------------------------- #
class MyRGBColor(Serialisable):
    tagname = 'srgbClr'
    namespace = DRAWING_NS
    val = Typed(expected_type=unicode)
    alpha = NestedInteger(allow_none=True)
    __elements__ = ('alpha',)

    def __init__(self, val, alpha=None):
        self.val = unicode(val)
        self.alpha = alpha


class MyColorChoice(ColorChoice):
    srgbClr = Typed(expected_type=MyRGBColor, allow_none=True)


# --------------------------------------------------------------- #
### END OF CLASS DEFINITIONS


### START OF MAIN PROGRAM
# Parsing the given arguments
options = parse_args()

# creating the Excel Workbook ...
workbook = Workbook()

# Creating the README sheet
mk_readme_sheet(workbook)

# Creating the Relative CPU usage Sheet
mk_cpurelative_sheet(workbook, options)

# Creating the Relative CPU usage Sheet
mk_cpuabsolute_sheet(workbook, options)

# saving the workbook
workbook.save(options['workbookname'])
workbook.close()
print ('{} created successfully'.format(options['workbookname']))
### END OF MAIN PROGRAM
